from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import List, Sequence
from dataclasses import asdict

from src.ingest.tabular_models import ColumnProfile, TabularProfile

import datetime as dt
import re
from typing import List, Sequence

def _normalize_header_name(name: str) -> str:
    """Normalize a CSV header to a stable, human-friendly identifier."""
    normalized = str(name).lstrip("\ufeff").strip()
    return " ".join(normalized.split())


def _normalize_cell_value(value: object) -> str:
    """Normalize a CSV cell by removing surrounding whitespace and repeated internal spacing."""
    if value is None:
        return ""
    normalized = str(value).strip()
    return " ".join(normalized.split()) if normalized else ""


_NUMERIC_PATTERN = re.compile(r"^-?\d+(\.\d+)?$")

def _is_iso_datetime(value: str) -> bool:
    try:
        dt.datetime.fromisoformat(value)
        return True
    except ValueError:
        try:
            dt.date.fromisoformat(value)
            return True
        except ValueError:
            return False

def _is_oca_numeric(value: str) -> bool:
    """True only for plain numeric values."""
    return bool(_NUMERIC_PATTERN.fullmatch(value))


def infer_datatype_candidates(values: Sequence[str]) -> List[str]:
    """Infer OCA datatype candidates for a column."""

    cleaned = [
        str(value).strip()
        for value in values
            if value is not None and str(value).strip()
    ]

    if not cleaned:
        return ["Text"]

    # Boolean
    if all(value.lower() in {"true", "false"} for value in cleaned):
        return ["Boolean"]

    # DateTime (strict ISO parsing)
    if all(_is_iso_datetime(value) for value in cleaned):
        return ["DateTime", "Text"]

    # Numeric (strict OCA numeric definition)
    if all(_is_oca_numeric(value) for value in cleaned):
        return ["Numeric"]

    return ["Text"]


def _is_iso_datetime(value: str) -> bool:
    try:
        dt.datetime.fromisoformat(value)
        return True
    except ValueError:
        return False


def profile_csv(csv_path: str | Path, output_path: str | Path | None = None) -> TabularProfile:
    path = Path(csv_path)

    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV file is empty: {path}")

    header = [_normalize_header_name(column) for column in rows[0]]
    data_rows = rows[1:]

    columns: List[ColumnProfile] = []
    for position, column_name in enumerate(header):
        values = [
            _normalize_cell_value(row[position]) if position < len(row) else ""
            for row in data_rows
        ]
        sample_values = values[:5]
        missing_count = sum(1 for value in values if value == "")
        unique_values = sorted({value for value in values if value != ""})

        column_profile = ColumnProfile(
            column_name=column_name,
            source_position=position,
            sample_values=sample_values,
            missing_count=missing_count,
            unique_count=len(unique_values),
            inferred_datatypes=infer_datatype_candidates(values),
        )
        columns.append(column_profile)

    profile = TabularProfile(source_file=str(path), columns=columns)

    if output_path is not None:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with output.open("w", encoding="utf-8") as handle:
            json.dump(asdict(profile), handle, indent=2)

    return profile


def _profile_rows(header: List[str], data_rows: List[List[str]], source_file: str) -> TabularProfile:
    columns: List[ColumnProfile] = []
    for position, column_name in enumerate(header):
        values = [
            _normalize_cell_value(row[position]) if position < len(row) else ""
            for row in data_rows
        ]
        sample_values = values[:5]
        missing_count = sum(1 for value in values if value == "")
        unique_values = sorted({value for value in values if value != ""})

        columns.append(
            ColumnProfile(
                column_name=column_name,
                source_position=position,
                sample_values=sample_values,
                missing_count=missing_count,
                unique_count=len(unique_values),
                inferred_datatypes=infer_datatype_candidates(values),
            )
        )

    return TabularProfile(source_file=source_file, columns=columns)


def profile_excel(
    excel_path: str | Path,
    output_path: str | Path | None = None,
    sheet_name: str | None = None,
) -> TabularProfile:
    from openpyxl import load_workbook

    path = Path(excel_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = list(worksheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        raise ValueError(f"Excel file is empty: {path}")

    header = [_normalize_header_name(str(column)) for column in rows[0]]
    data_rows = [
        [_normalize_cell_value(value) for value in row]
        for row in rows[1:]
    ]

    profile = _profile_rows(header, data_rows, str(path))

    if output_path is not None:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with output.open("w", encoding="utf-8") as handle:
            json.dump(asdict(profile), handle, indent=2)

    return profile


def profile_tabular(path: str | Path, output_path: str | Path | None = None) -> TabularProfile:
    tabular_path = Path(path)
    suffix = tabular_path.suffix.lower()
    if suffix == ".csv":
        return profile_csv(tabular_path, output_path)
    if suffix in {".xlsx", ".xlsm"}:
        return profile_excel(tabular_path, output_path)
    raise ValueError(f"Unsupported tabular file type: {tabular_path}")
